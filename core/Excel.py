import datetime

from openpyxl import load_workbook

from core.Mssql import Mssql


class Excel:
    def __init__(self, file_path, target, incremental=False):
        self.file_path = file_path
        self.target = target
        self.incremental = incremental
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.columns_names = self.get_column_names()
        self.row_count = self.get_row_count()
        self.fields_map = {}
        self.filters = []
        self.filters_tuple = []

        self.merge_count = 0
        self.drop_count = 0

        self.tag = None

    def __str__(self):
        return 'Excel: ' + self.file_path

    def get_column_names(self):
        column_num = 1
        column_names = []
        while True:
            value = self.ws.cell(1, column_num).value
            if value is not None:
                column_names.append(value)
                column_num += 1
            else:
                break
        return column_names

    def get_row_count(self):
        row_num = 2
        row_count = 0
        while True:
            value = self.ws.cell(row_num, 1).value
            if value is not None:
                row_count += 1
                row_num += 1
            else:
                break
        return row_count

    def add_map(self, source_name, target_name):
        if source_name not in self.columns_names:
            return False
        if target_name not in self.target.get_current_table_detail():
            return False
        if target_name in self.fields_map:
            return False
        self.fields_map[target_name] = source_name
        return True

    def add_filter(self, source_name, filter_type, value):
        if filter_type == 'gt':
            self.filters.append(lambda datadict: datadict[source_name] > value)
        elif filter_type == 'ge':
            self.filters.append(lambda datadict: datadict[source_name] >= value)
        elif filter_type == 'lt':
            self.filters.append(lambda datadict: datadict[source_name] < value)
        elif filter_type == 'le':
            self.filters.append(lambda datadict: datadict[source_name] <= value)
        elif filter_type == 'eq':
            self.filters.append(lambda datadict: datadict[source_name] == value)
        elif filter_type == 'contain':
            self.filters.append(lambda datadict: value in datadict[source_name])
        elif filter_type == 'notcontain':
            self.filters.append(lambda datadict: value not in datadict[source_name])

    def apply_filter(self, datadict):
        for filter_ in self.filters:
            if not filter_(datadict):
                return False
        return True

    def merge_to_target(self):
        merge_count = 0
        drop_count = 0
        max_col = len(self.columns_names)
        if self.incremental:
            max_col += 1
        for row in self.ws.iter_rows(min_row=2,
                                     max_col=max_col,
                                     max_row=2 + self.row_count - 1):
            table = self.target.table
            ins = table.insert()
            source_data = [(source_name, row[self.columns_names.index(source_name)].value)
                           for source_name in self.columns_names]
            source_datadict = dict(source_data)
            if not self.apply_filter(source_datadict):
                continue
            if self.incremental:
                if row[len(self.columns_names)].value:
                    drop_count += 1
                    continue
                else:
                    row[len(self.columns_names)].value = True
            data = [(target_name, row[self.columns_names.index(source_name)].value)
                    for target_name, source_name in self.fields_map.items()]
            datadict = dict(data)
            self.target.conn.execute(ins, datadict)
            merge_count += 1
        self.wb.save(self.file_path)
        self.merge_count = merge_count
        self.drop_count = drop_count


if __name__ == '__main__':
    target = Mssql('sa', '132132qq', 'a')
    target.connect_db()
    target.set_table('fake_data01')
    excel = Excel('data.xlsx', target, incremental=True)
    print(excel.columns_names)
    print(excel.row_count)
    excel.add_map('name_excel', 'name1')
    excel.add_map('age_excel', 'age1')
    excel.add_map('salary_excel', 'salary1')
    excel.add_map('birthday_excel', 'birthday1')
    excel.add_map('is_human_excel', 'is_human1')
    excel.add_filter('id', 'gt', 50)
    excel.add_filter('birthday_excel', 'gt', datetime.datetime(2000, 1, 1))
    print(excel.fields_map)
    excel.merge_to_target()
