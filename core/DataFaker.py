import datetime
import os
import random
import pymongo
from openpyxl import Workbook, load_workbook
from core.Mongodb import Mongodb
from core.Mssql import Mssql
from faker import Faker
from core.Mysql import Mysql
from config import config_mssql, config_mysql, config_excel, config_mongodb

fake = Faker()


class MssqlFaker:
    def __init__(self, username, password, dsn):
        self.mssql = Mssql(username, password, dsn)
        self.mssql.connect_db()

    def gen_fake_data(self, table_name, field_suffix, insert_count=100):
        self.mssql.connect_db()
        fields_pair = [['name', 'string'],
                       ['age', 'int'],
                       ['salary', 'float'],
                       ['birthday', 'datetime'],
                       ['is_human', 'bool']]
        for pair in fields_pair:
            pair[0] += ('_' + field_suffix)
        print(fields_pair)
        fields = dict(fields_pair)

        table = self.mssql.create_table(table_name, fields)
        ins = table.insert()

        self.mssql.conn.execute(
            ins, [{
                fields_pair[0][0]: fake.name(),
                fields_pair[1][0]: random.randint(1, 89),
                fields_pair[2][0]: random.randint(10000, 20000) / 100,
                fields_pair[3][0]: datetime.datetime(random.randint(1980, 2010, ),
                                                     random.randint(1, 12),
                                                     random.randint(1, 28)),
                fields_pair[4][0]: True if random.randint(0, 1) == 1 else False
            } for _ in range(insert_count)]
        )


class ExcelFaker:
    def __init__(self, file_path):
        self.file_path = file_path

    def gen_fake_data(self, insert_count=100):
        wb = Workbook()
        ws = wb.active
        ws.title = '学生信息表'
        ws['A1'] = 'id'
        ws['B1'] = 'name_excel'
        ws['C1'] = 'age_excel'
        ws['D1'] = 'salary_excel'
        ws['E1'] = 'birthday_excel'
        ws['F1'] = 'is_human_excel'
        count = 1
        fake = Faker()
        for row in ws.iter_rows(min_row=2,
                                max_col=6,
                                max_row=2 + insert_count - 1):
            row[0].value = count
            count += 1
            row[1].value = fake.name()
            row[2].value = random.randint(1, 55)
            row[3].value = random.randint(10000, 99999) / 100
            row[4].value = datetime.date(
                random.randint(1970, 2019),
                random.randint(1, 12),
                random.randint(1, 28)
            )
            row[5].value = True if random.randint(0, 1) > 0 else False
        else:
            wb.save(self.file_path)

    def append_fake_data(self, insert_count=100):
        wb = load_workbook(self.file_path)
        ws = wb.active
        row_num = 2
        last_id = 0
        while True:
            value = ws.cell(row_num, 1).value
            if value is not None:
                last_id = value
                row_num += 1
            else:
                break
        current_id = last_id + 1
        fake = Faker()
        for row in ws.iter_rows(min_row=row_num,
                                max_col=6,
                                max_row=row_num + insert_count - 1):
            row[0].value = current_id
            current_id += 1
            row[1].value = fake.name()
            row[2].value = random.randint(1, 55)
            row[3].value = random.randint(10000, 99999) / 100
            row[4].value = datetime.date(
                random.randint(1970, 2019),
                random.randint(1, 12),
                random.randint(1, 28)
            )
            row[5].value = True if random.randint(0, 1) > 0 else False
        wb.save(self.file_path)


class MysqlFaker:
    def __init__(self, username, password, host, db):
        self.mysql = Mysql(username, password, host, db)
        self.mysql.connect_db()

    def gen_fake_data(self, table_name, field_suffix, insert_count=100):
        self.mysql.connect_db()
        fields_pair = [['name', 'string'],
                       ['age', 'int'],
                       ['salary', 'float'],
                       ['birthday', 'datetime'],
                       ['is_human', 'bool']]
        for pair in fields_pair:
            pair[0] += ('_' + field_suffix)
        print(fields_pair)
        fields = dict(fields_pair)

        table = self.mysql.create_table(table_name, fields)
        ins = table.insert()

        self.mysql.conn.execute(
            ins, [{
                fields_pair[0][0]: fake.name(),
                fields_pair[1][0]: random.randint(1, 89),
                fields_pair[2][0]: random.randint(10000, 20000) / 100,
                fields_pair[3][0]: datetime.datetime(random.randint(1980, 2010, ),
                                                     random.randint(1, 12),
                                                     random.randint(1, 28)),
                fields_pair[4][0]: True if random.randint(0, 1) == 1 else False
            } for _ in range(insert_count)])


class MongodbFaker:
    def __init__(self, db='fake', collection='fake_mongo'):
        self.mongodb = Mongodb(db, collection, None)

    def gen_fake_data(self, collection_name='fake_mongo',
                      insert_count=100):
        collection = self.mongodb.db[collection_name]
        for i in range(insert_count):
            collection.insert_one(
                {'id': i,
                 'name_mongo': fake.name(),
                 'age_mongo': random.randint(1, 99),
                 'salary_mongo': random.randint(10000, 20000) / 100,
                 'birthday_mongo': datetime.datetime(random.randint(1980, 2010),
                                                     random.randint(1, 12),
                                                     random.randint(1, 28)),
                 'is_human_mongo': True if random.randint(0, 1) > 0 else False}
            )


def create():
    mssql_faker = MssqlFaker(config_mssql['username'],
                             config_mssql['password'],
                             config_mssql['dsn'])
    mssql_faker.gen_fake_data(config_mssql['fake_table_name'], 'mssql')
    mssql_faker.gen_fake_data(config_mssql['fake_target_name'], 'mssql')

    mysql_faker = MysqlFaker(config_mysql['username'],
                             config_mysql['password'],
                             config_mysql['host'],
                             config_mysql['db'])
    mysql_faker.gen_fake_data(config_mysql['fake_table_name'], 'mysql')

    excel_faker = ExcelFaker(file_path=config_excel['file_path'])
    excel_faker.gen_fake_data()

    mongo_faker = MongodbFaker(config_mongodb['fake_db'])
    mongo_faker.gen_fake_data(config_mongodb['fake_collection'])


def drop():
    mssql = Mssql(config_mssql['username'],
                  config_mssql['password'],
                  config_mssql['dsn'])
    mssql.connect_db()
    mssql.drop_table(config_mssql['fake_table_name'])
    mssql.drop_table(config_mssql['fake_target_name'])
    mssql.drop_table(config_mssql['fake_table_name'] + '_to_'
                     + config_mssql['fake_target_name'])

    mysql = Mysql(config_mysql['username'],
                  config_mysql['password'],
                  config_mysql['host'],
                  config_mysql['db'])
    mysql.connect_db()
    mysql.drop_table(config_mysql['fake_table_name'])
    mysql.drop_table(config_mysql['fake_table_name'] + '_to_'
                     + config_mssql['fake_target_name'])
    if os.path.exists(config_excel['file_path']):
        os.remove(config_excel['file_path'])

    mongo = pymongo.MongoClient()
    mongo.drop_database(config_mongodb['fake_db'])


if __name__ == '__main__':
    drop()
    create()
